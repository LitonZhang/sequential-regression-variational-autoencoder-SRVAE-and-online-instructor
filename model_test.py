from collections import defaultdict
import numpy as np
import openpyxl
import torch.utils.data as Data
import LSTMTrain
from SRVAE import *
import argparse
from deep_factors import *
import util
import os


def test_data(dataset_x, dataset_y, time_step):
    data_x = []
    data_y = []
    num_row = dataset_x.shape[1]
    for i in range(0, num_row - time_step):
        data_x.append([a for a in dataset_x[0, i:i + time_step]])
        data_y.append([a for a in dataset_y[0, i:i + time_step]])
    data_x, data_y = np.array(data_x), np.array(data_y)
    data_x, data_y = torch.tensor(data_x).float(), torch.tensor(data_y).float()
    return data_x, data_y


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--pre_train', "-p", action='store_true', default=False)  # 预训练
    parser.add_argument('--online_train', "-o", action='store_true', default=False)  # DQN训练
    parser.add_argument("--modelname", type=str, default='SRVAE')
    parser.add_argument("--processID", type=int, default=1)
    parser.add_argument("--pre_epoch_nums", type=int, default=400)  # 1000 保证多个模型的泛化性能
    parser.add_argument("--pre_batch_size", type=int, default=512)  # 256
    parser.add_argument("--n_factors", type=int, default=14)  # 14
    parser.add_argument("--global_hidden_size", type=int, default=256)  # 128  提取全局信息0
    parser.add_argument("--global_nlayers", type=int, default=1)  # 2
    parser.add_argument("--timestep", type=int, default=6)  # 2
    parser.add_argument("-Adam_lr", type=float, default=7e-3)  # 大部分4e-3，少部分5e-3
    parser.add_argument("--multiple", type=int, default=3)  # 1,2,3
    parser.add_argument("--Monte_size", type=int, default=20)  # 100
    # online
    parser.add_argument("--online_epoch_nums", type=int, default=400)  # 1000
    parser.add_argument("--online_batch_size", type=int, default=5)
    parser.add_argument("-ftml_lr", type=float, default=1e-3)  # 1e-5
    parser.add_argument("--test_length", type=int, default=30)
    # DQN
    parser.add_argument('--DQN_batch_size', type=int, default=32)
    parser.add_argument('--DQN_LR', type=float, default=1e-4)
    parser.add_argument('--DQNepsilon', type=float, default=0.8)
    parser.add_argument('--DQN_GAMMA', type=float, default=0.9)
    parser.add_argument('--memory_capacity', type=int, default=1000, help='do not use scientific notation')  # 10000
    parser.add_argument('--DQN_target_update', type=int, default=6)
    # function
    parser.add_argument('--n_trials', type=int, default=100)
    parser.add_argument("--show_plot", "-sp", action="store_true")
    parser.add_argument('--no_cuda', action='store_true', default=False, help='True to disable CUDA')
    parser.add_argument('--seed', type=int, default=1, metavar='S', help='random seed (default: 1)')
    # 封装args
    args = parser.parse_args()
    # GPU加速
    args.cuda = not args.no_cuda and torch.cuda.is_available()  # 判断pytorch是否有GPU加速
    # print(torch.cuda.is_available())
    torch.manual_seed(args.seed)  # 为所有GPU设置种子数
    device = torch.device("cuda" if args.cuda else "cpu")  # 代表将torch.Tensor分配到的设备的对象
    print('Start running on {}'.format(device))
    kwargs = {'num_workers': 1, 'pin_memory': True} if args.cuda else {}

    # 批量创建DQN模型
    DQN_list = defaultdict(list)
    for processID in range(1, 2):
        DQN_list[processID - 1] = torch.load(
            os.path.join('dqn_models', 'DQNp' + str(processID) + 'e' + str(args.online_epoch_nums) + '.pt'),
            map_location=device)
    # 模型实例化管理器
    model_manager = util.ModelManager(args, device)

    # 批量创建预测模型
    net = defaultdict(list)
    pre_net = defaultdict(list)
    # 加载在线学习模型
    for processID in range(1, args.processID + 1):
        # 批量实例化模型
        net[args.modelname].append(
            SRVAE(4, args.global_nlayers, args.global_hidden_size, args.n_factors, args.multiple).to(device))
        # 加载模型参数
        model_manager.load_models(net[args.modelname][processID - 1], processID, 'online_saved_models')
    # 加载预训练模型
    for processID in range(1, args.processID + 1):
        # 批量实例化模型
        pre_net[args.modelname].append(
            SRVAE(4, args.global_nlayers, args.global_hidden_size, args.n_factors, args.multiple).to(device))
        # 加载模型参数
        model_manager.load_models(pre_net[args.modelname][processID - 1], processID, 'saved_models')
    online_model = net[args.modelname][args.processID - 1]
    pre_model = pre_net[args.modelname][args.processID - 1]
    # 加载LSTM模型
    LSTM_model = torch.load(os.path.join('LSTM_model', 'p' + str(args.processID) + '.pt'))
    # 加载DeepFactors模型
    dp_model = torch.load(os.path.join('DF_model', 'p' + str(args.processID) + '.pt'))

    # 选择归一化方法
    x_scaler = util.MeanScaler()
    y_scaler = util.MeanScaler()

    # 打开以下注释，将所有数据进行测试
    # 数据提取
    # period = [2018, 1, 1, 2022, 6, 8]  # 最后一个单子:2021, 12, 20 最后一行:2022, 5, 8 第一行:2018, 1, 1  real-test:2022,2 27
    # raw_x, raw_y = util.extract_data('data.csv', period, process=args.processID)
    # # 数据划分
    # test_x, test_y = test_data(raw_x, raw_y, args.timestep)
    # # 数据
    # _, mean_x = x_scaler.fit_transform(test_y)
    # test_y, mean_y = y_scaler.fit_transform(test_y)
    # # 打包测试集
    # torch_test_data = Data.TensorDataset(test_x, test_y)
    # test_loader = Data.DataLoader(dataset=torch_test_data, batch_size=args.online_batch_size, shuffle=False)

    # 打开以下注释，仅对测试集数据进行测试
    period = [2018, 1, 1, 2022, 6, 8]
    pre_train_loader, online_train_loader, test_loader, dim_0, num_rows, num_features, mean_x, mean_y, online_train_y = util.to_data_loader(
        args, period)

    # testing
    test_loss, on_results, reals = test(online_model, test_loader, device, args)
    pre_loss, pre_results, _ = test(pre_model, test_loader, device, args)
    dp_loss, dp_results, _ = dftest(dp_model, test_loader, args, device='cpu')
    LSTM_loss, LSTM_results, _ = LSTMTrain.test(LSTM_model, test_loader, device, args)

    reals = reals * mean_y
    on_results = on_results * mean_y
    pre_results = pre_results * mean_y
    dp_results = dp_results * mean_y
    LSTM_results = LSTM_results * mean_y

    # figure
    real_plot = reals.cpu().numpy()[-args.test_length:, 0].reshape(-1)
    pre_plot = pre_results.cpu().numpy()[-args.test_length:, 0].reshape(-1)
    on_plot = on_results.cpu().numpy()[-args.test_length:, 0].reshape(-1)
    dp_plot = dp_results.cpu().numpy()[-args.test_length:, 0].reshape(-1)
    LSTM_plot = LSTM_results.cpu().numpy()[-args.test_length:, 0].reshape(-1)

    plt.plot(real_plot, c='black', label='real', linewidth=1.2)
    plt.plot(pre_plot, c='blue', label='SRVAE_pretrain', linewidth=1.2)
    plt.plot(on_plot, c='green', label='SRVAE_online_train', linewidth=1.2)  # 显示多少个预测值
    plt.plot(dp_plot, c='#d1b26f', label='Deep_factors', linewidth=1.2)  # 显示多少个预测值
    plt.plot(LSTM_plot, c='darkorange', label='LSTM', linewidth=1.2)  # 显示多少个预测值

    plt.legend()
    plt.show()

    try:
        EXCEL = openpyxl.load_workbook('result/test.xlsx')  # 存在就打开test的xlsx
    except Exception as e:
        EXCEL = openpyxl.Workbook()  # 不存在就创建
    sheet_name = EXCEL.active  # 获取当前活跃的sheet,默认是第一个sheet
    sheet_name.delete_rows(1,sheet_name.max_row)
    # 写入表头信息
    my_title = ['Real value', 'SRVAE online', 'SRVAE offline', 'Deep AR', 'LSTM']
    sheet_name.append(my_title)
    print("Data写入EXCEL")
    # 把表头写入第一行
    for i in range(5):  # 一共5列数据
        if i == 0:  # 第0列
            for j, value in enumerate(reals[:, 0].cpu().numpy()):  # excel中行列都是从1开始
                sheet_name.cell(row=j + 2, column=i + 1, value=value)
        if i == 2:  # 第1列
            for m, value in enumerate(pre_results[:, 0].cpu().numpy()):
                sheet_name.cell(row=m + 2, column=i + 1, value=value)
        if i == 1:  # 第3列
            for n, value in enumerate(on_results[:, 0].cpu().numpy()):
                sheet_name.cell(row=n + 2, column=i + 1, value=value)
        if i == 3:  # 第2列
            for o, value in enumerate(dp_results[:, 0].cpu().numpy()):
                sheet_name.cell(row=o + 2, column=i + 1, value=value)
        if i == 4:  # 第4列
            for o, value in enumerate(LSTM_results[:, 0].cpu().numpy()):
                sheet_name.cell(row=o + 2, column=i + 1, value=value)
    EXCEL.save('result/test.xlsx')  # 相对路径保存工作簿
    EXCEL.close()
